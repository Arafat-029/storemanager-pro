from __future__ import annotations
from pathlib import Path
from datetime import datetime
import html
import os
import re
import tempfile
import uuid

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image as PILImage, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from config import RECEIPTS_DIR


_ARABIC_FONT_CACHE: str | None | bool = False


def export_to_excel(data: list[dict], filename: str, sheet_name: str = "Données") -> str:
    RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)
    path = RECEIPTS_DIR / f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        wb.save(str(path))
        return str(path)

    headers = list(data[0].keys())
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key))

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(str(path))
    return str(path)


def _contains_arabic(text: str | None) -> bool:
    if not text:
        return False
    return any(
        "\u0600" <= ch <= "\u06FF"
        or "\u0750" <= ch <= "\u077F"
        or "\u08A0" <= ch <= "\u08FF"
        or "\uFB50" <= ch <= "\uFDFF"
        or "\uFE70" <= ch <= "\uFEFF"
        for ch in str(text)
    )


def _find_arabic_font_path() -> str | None:
    global _ARABIC_FONT_CACHE
    if _ARABIC_FONT_CACHE is not False:
        return _ARABIC_FONT_CACHE or None

    candidates = [
        # Windows
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\arialbd.ttf",
        r"C:\Windows\Fonts\tahoma.ttf",
        r"C:\Windows\Fonts\tahomabd.ttf",
        r"C:\Windows\Fonts\segoeui.ttf",
        r"C:\Windows\Fonts\seguisb.ttf",
        # Linux
        "/usr/share/fonts/opentype/fonts-hosny-amiri/Amiri-Regular.ttf",
        "/usr/share/fonts/opentype/fonts-hosny-amiri/Amiri-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        # macOS
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode MS.ttf",
    ]
    for cand in candidates:
        if os.path.exists(cand):
            _ARABIC_FONT_CACHE = cand
            return cand

    _ARABIC_FONT_CACHE = None
    return None



def _arabic_forms(ch: str) -> tuple[str | None, str | None, str | None, str | None]:
    try:
        base_name = __import__("unicodedata").name(ch)
    except ValueError:
        return (None, None, None, None)

    def _lookup(suffix: str) -> str | None:
        try:
            return __import__("unicodedata").lookup(f"{base_name} {suffix}")
        except KeyError:
            return None

    return (
        _lookup("ISOLATED FORM"),
        _lookup("FINAL FORM"),
        _lookup("INITIAL FORM"),
        _lookup("MEDIAL FORM"),
    )


def _can_join_prev(ch: str) -> bool:
    isolated, final, initial, medial = _arabic_forms(ch)
    return bool(final or medial)


def _can_join_next(ch: str) -> bool:
    isolated, final, initial, medial = _arabic_forms(ch)
    return bool(initial or medial)


def _shape_arabic_word(word: str) -> str:
    chars = list(word)
    shaped: list[str] = []

    for i, ch in enumerate(chars):
        if not _contains_arabic(ch):
            shaped.append(ch)
            continue

        prev_ch = chars[i - 1] if i > 0 else ""
        next_ch = chars[i + 1] if i + 1 < len(chars) else ""

        joins_prev = bool(prev_ch and _contains_arabic(prev_ch) and _can_join_next(prev_ch) and _can_join_prev(ch))
        joins_next = bool(next_ch and _contains_arabic(next_ch) and _can_join_next(ch) and _can_join_prev(next_ch))

        isolated, final, initial, medial = _arabic_forms(ch)

        if joins_prev and joins_next and medial:
            shaped.append(medial)
        elif joins_prev and final:
            shaped.append(final)
        elif joins_next and initial:
            shaped.append(initial)
        else:
            shaped.append(isolated or ch)

    return "".join(reversed(shaped))


def _prepare_rtl_text(text: str) -> str:
    if not _contains_arabic(text):
        return text

    parts = re.split(r"(\s+)", text or "")
    visual_parts: list[str] = []

    for part in parts:
        if not part:
            continue
        if part.isspace():
            visual_parts.append(part)
        elif _contains_arabic(part):
            visual_parts.append(_shape_arabic_word(part))
        else:
            visual_parts.append(part)

    return "".join(reversed(visual_parts))

def _paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(html.escape(text or ""), style)


def _rtl_image_flowable(
    text: str,
    *,
    max_width: float,
    font_size: int,
    bold: bool = False,
    centered: bool = False,
    align: str = "left",
) -> RLImage | None:
    font_path = _find_arabic_font_path()
    if not font_path:
        return None

    visual_text = _prepare_rtl_text(text)
    cache_dir = RECEIPTS_DIR / ".receipt_rtl_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)

    scale = 2
    pad_x = 8 * scale
    pad_y = 4 * scale
    max_width_px = max(40, int(max_width * scale))

    chosen_size = max(8, font_size * scale)
    font = None
    for size in range(chosen_size, 11, -1):
        try:
            probe_font = ImageFont.truetype(font_path, size=size)
        except Exception:
            return None
        probe_img = PILImage.new("RGBA", (max_width_px, size * 3), (255, 255, 255, 0))
        probe_draw = ImageDraw.Draw(probe_img)
        bbox = probe_draw.textbbox((0, 0), visual_text, font=probe_font)
        text_w = max(1, bbox[2] - bbox[0])
        if text_w <= max_width_px - (2 * pad_x):
            font = probe_font
            chosen_size = size
            break

    if font is None:
        font = ImageFont.truetype(font_path, size=12)

    tmp = PILImage.new("RGBA", (max_width_px, chosen_size * 4), (255, 255, 255, 0))
    draw = ImageDraw.Draw(tmp)
    bbox = draw.textbbox((0, 0), visual_text, font=font)

    text_w = max(1, bbox[2] - bbox[0])
    text_h = max(1, bbox[3] - bbox[1])
    img_w = min(max_width_px, text_w + (2 * pad_x))
    img_h = text_h + (2 * pad_y)
    img = PILImage.new("RGBA", (img_w, img_h), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    if centered:
        x = max(pad_x, (img_w - text_w) // 2)
    elif align == "right":
        x = max(pad_x, img_w - text_w - pad_x)
    else:
        x = pad_x

    y = pad_y - bbox[1]
    draw.text((x, y), visual_text, font=font, fill="black")

    filename = cache_dir / f"rtl_{uuid.uuid4().hex}.png"
    img.save(filename)
    flow = RLImage(str(filename), width=img_w / scale, height=img_h / scale)
    flow.hAlign = "LEFT" if align != "right" else "RIGHT"
    return flow


def _receipt_text_flowable(
    text: str,
    style: ParagraphStyle,
    *,
    max_width: float,
    font_size: int | None = None,
    bold: bool = False,
    centered: bool = False,
):
    if _contains_arabic(text):
        flow = _rtl_image_flowable(
            text,
            max_width=max_width,
            font_size=font_size or int(style.fontSize),
            bold=bold,
            centered=centered,
        )
        if flow is not None:
            return flow
    return _paragraph(text, style)


def _item_name_cell(name: str, width: float, font_size: int):
    if _contains_arabic(name):
        flow = _rtl_image_flowable(name, max_width=width - 4, font_size=font_size, centered=False, align="left")
        if flow is not None:
            return flow
    return name


def generate_receipt_pdf(sale: dict, store_info: dict) -> str:
    RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"ticket_{sale['id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    path = RECEIPTS_DIR / filename

    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=30, leftMargin=30, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle("title_store", parent=styles["Title"], alignment=TA_CENTER)
    normal_center = ParagraphStyle("normal_center", parent=styles["Normal"], alignment=TA_CENTER)
    normal_left = ParagraphStyle("normal_left", parent=styles["Normal"], alignment=TA_LEFT)

    story.append(_receipt_text_flowable(
        store_info.get("store_name", "Magasin"),
        title_style,
        max_width=A4[0] - 60,
        font_size=18,
        bold=True,
        centered=True,
    ))
    address = store_info.get("store_address", "")
    if address:
        story.append(_receipt_text_flowable(address, normal_center, max_width=A4[0] - 60, font_size=10, centered=True))
    phone = store_info.get("store_phone", "")
    if phone:
        story.append(_paragraph(f"Tél: {phone}", normal_center))
    story.append(Spacer(1, 12))
    story.append(_paragraph(f"Ticket N°: {sale['id']}", normal_left))
    story.append(_paragraph(f"Date: {sale['created_at']}", normal_left))
    story.append(_paragraph(f"Caissier: {sale.get('cashier_name', '')}", normal_left))
    story.append(Spacer(1, 12))

    table_data = [["Produit", "Qté", "P.U", "Total"]]
    name_col_width = 200
    for item in sale.get("items", []):
        product_name = str(item["product_name"])
        if item.get("sale_unit_name"):
            product_name = f"{product_name} [{item['sale_unit_name']}]"
        qty_value = float(item.get("quantity") or 0.0)
        qty_text = f"{int(round(qty_value))}" if abs(qty_value - round(qty_value)) < 0.001 else f"{qty_value:.2f}"
        table_data.append([
            _item_name_cell(product_name, name_col_width, 10),
            qty_text,
            f"{item['unit_price']:.3f}",
            f"{item['total']:.3f}",
        ])

    t = Table(table_data, colWidths=[name_col_width, 60, 80, 80])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))

    currency = store_info.get("currency", "TND")
    story.append(_paragraph(f"Sous-total: {sale['subtotal']:.3f} {currency}", styles["Normal"]))
    if sale.get("discount", 0):
        story.append(_paragraph(f"Remise: -{sale['discount']:.3f} {currency}", styles["Normal"]))
    if sale.get("tax", 0):
        story.append(_paragraph(f"TVA: {sale['tax']:.3f} {currency}", styles["Normal"]))
    story.append(_paragraph(f"TOTAL: {sale['total']:.3f} {currency}", styles["Heading2"]))
    story.append(_paragraph(f"Payé: {sale['amount_paid']:.3f} {currency}", styles["Normal"]))
    story.append(_paragraph(f"Monnaie: {sale['change_given']:.3f} {currency}", styles["Normal"]))
    story.append(Spacer(1, 20))
    story.append(_receipt_text_flowable(
        store_info.get("receipt_footer", "Merci !"),
        normal_center,
        max_width=A4[0] - 60,
        font_size=10,
        centered=True,
    ))

    doc.build(story)
    return str(path)


def generate_thermal_receipt(sale: dict, store_info: dict) -> str:
    """Generate a narrow receipt PDF sized for an 80 mm thermal / cash-register printer."""
    RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"ticket_{sale['id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    path = RECEIPTS_DIR / filename

    PW     = 226.77
    MARGIN = 8.0
    CW     = PW - 2 * MARGIN

    hdr = ParagraphStyle("hdr", fontName="Helvetica-Bold", fontSize=11, alignment=TA_CENTER, spaceAfter=1, leading=14)
    sml = ParagraphStyle("sml", fontName="Helvetica", fontSize=8, alignment=TA_CENTER, spaceAfter=1, leading=10)
    nrm = ParagraphStyle("nrm", fontName="Helvetica", fontSize=8, alignment=TA_LEFT, spaceAfter=1, leading=10)
    tot = ParagraphStyle("tot", fontName="Helvetica-Bold", fontSize=11, alignment=TA_LEFT, spaceAfter=2, leading=14)

    DIV  = "─" * 34
    DDIV = "═" * 34

    currency = store_info.get("currency", "TND")
    story: list = []

    story.append(_paragraph(DDIV, sml))
    story.append(_receipt_text_flowable(
        store_info.get("store_name", "Magasin"),
        hdr,
        max_width=CW,
        font_size=11,
        bold=True,
        centered=True,
    ))
    addr = store_info.get("store_address") or store_info.get("address", "")
    if addr:
        story.append(_receipt_text_flowable(addr, sml, max_width=CW, font_size=8, centered=True))
    phone = store_info.get("store_phone") or store_info.get("phone", "")
    if phone:
        story.append(_paragraph(f"Tél : {phone}", sml))
    story.append(_paragraph(DDIV, sml))
    story.append(Spacer(1, 4))

    story.append(_paragraph(f"Ticket N° : {sale['id']}", nrm))
    story.append(_paragraph(f"Date : {sale['created_at']}", nrm))
    if sale.get("cashier_name"):
        story.append(_paragraph(f"Caissier : {sale['cashier_name']}", nrm))
    story.append(Spacer(1, 4))
    story.append(_paragraph(DIV, sml))

    col_w = [CW - 100, 28, 36, 36]
    rows = [["Produit", "Qté", "P.U.", "Total"]]
    for item in sale.get("items", []):
        name = str(item["product_name"])
        if item.get("sale_unit_name"):
            name = f"{name} [{item['sale_unit_name']}]"
        if not _contains_arabic(name) and len(name) > 17:
            name = name[:16] + "…"
        qty_value = float(item.get("quantity") or 0.0)
        qty_text = f"{int(round(qty_value))}" if abs(qty_value - round(qty_value)) < 0.001 else f"{qty_value:.2f}"
        rows.append([
            _item_name_cell(name, col_w[0], 9),
            qty_text,
            f"{item['unit_price']:.3f}",
            f"{item['total']:.3f}",
        ])
    tbl = Table(rows, colWidths=col_w)
    tbl.setStyle(TableStyle([
        ("FONTNAME",       (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",       (0, 0), (-1, -1), 7),
        ("LINEBELOW",      (0, 0), (-1,  0), 0.4, colors.grey),
        ("ALIGN",          (0, 1), (0, -1), "LEFT"),
        ("ALIGN",          (1, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING",     (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 2),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 4))
    story.append(_paragraph(DIV, sml))

    story.append(_paragraph(f"Sous-total : {sale['subtotal']:.3f} {currency}", nrm))
    if sale.get("discount", 0):
        story.append(_paragraph(f"Remise :    -{sale['discount']:.3f} {currency}", nrm))
    if sale.get("tax", 0):
        story.append(_paragraph(f"TVA :        {sale['tax']:.3f} {currency}", nrm))
    story.append(_paragraph(DDIV, sml))
    story.append(_paragraph(f"TOTAL : {sale['total']:.3f} {currency}", tot))
    story.append(_paragraph(DDIV, sml))
    story.append(_paragraph(f"Payé :    {sale['amount_paid']:.3f} {currency}", nrm))
    story.append(_paragraph(f"Monnaie : {sale['change_given']:.3f} {currency}", nrm))
    story.append(Spacer(1, 8))
    story.append(_receipt_text_flowable(
        store_info.get("receipt_footer", "Merci de votre visite !"),
        sml,
        max_width=CW,
        font_size=8,
        centered=True,
    ))
    story.append(_paragraph(DDIV, sml))

    doc = SimpleDocTemplate(
        str(path),
        pagesize=(PW, A4[1]),
        rightMargin=MARGIN,
        leftMargin=MARGIN,
        topMargin=MARGIN,
        bottomMargin=MARGIN,
    )
    doc.build(story)
    return str(path)
